import openpyxl, os
#import PyPDF2  --> For pdf file operations
#import docx --> For document file operations

#os.chdir('C:\\users\\all\documents') #change dir

#Reading
workbook = openpyxl.load_workbook('example.xlsx')
type(workbook)

sheet = workbook.get_sheet_by_name('Sheet1')
cellobj = sheet['A1']
val = cellobj.value

sheet.cell(row=1, column=2)

# multiple rows in col=2
for i in range(1,8):
    print(i,sheet.cell(row=i, column=2).value)

#Editing
wb = openpyxl.Workbook()
wb.get_sheet_names()
sh = wb.get_sheet_by_name('Sheet')
sh['A1'] = 34
sh['A2'] = 'voila!'
wb.save('example.xslx')